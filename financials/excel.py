"""Excel workbook builder — all sheet-building functions + build_full_workbook()."""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .formatters import fmt_money, fmt_val

# ── Colours ──────────────────────────────────────────────────────────────────
DARK_BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "F0F4F8"
WHITE = "FFFFFF"
GREEN = "00703C"
RED = "CC0000"
CELL_GREEN = "C6EFCE"
CELL_RED = "FFCCCC"
CELL_NEUT = "F5F5F5"
TEXT_GREEN = "276221"
TEXT_RED = "9C0006"


def header_style(cell, bg=DARK_BLUE, fg=WHITE):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def alt_fill(cell):
    cell.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")


def title_style(cell, size=14):
    cell.font = Font(bold=True, size=size, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _safe_avg(values):
    clean = [v for v in values if v is not None and not (isinstance(v, float) and pd.isna(v))]
    return sum(clean) / len(clean) if clean else None


def _cmp_color(company_val, avg_val, higher_is_better: bool, threshold: float = 0.05):
    if company_val is None or avg_val is None or avg_val == 0:
        return CELL_NEUT, "000000"
    diff = (company_val - avg_val) / abs(avg_val)
    if abs(diff) <= threshold:
        return CELL_NEUT, "000000"
    beats = diff > 0 if higher_is_better else diff < 0
    return (CELL_GREEN, TEXT_GREEN) if beats else (CELL_RED, TEXT_RED)


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_overview_sheet(ws, info: dict):
    ws.merge_cells("A1:C1")
    title_style(ws["A1"])
    company = info.get("longName", "N/A")
    symbol = info.get("symbol", "")
    ws["A1"].value = f"{company} ({symbol})  --  Financial Overview"
    ws.row_dimensions[1].height = 32

    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws.row_dimensions[2].height = 14

    rows = [
        ("Metric", "Value"),
        ("Company", company),
        ("Sector", info.get("sector", "N/A")),
        ("Industry", info.get("industry", "N/A")),
        ("Market Cap", fmt_money(info.get("marketCap"))),
        ("Current Price", fmt_val(info.get("currentPrice") or info.get("regularMarketPrice"), prefix="$")),
        ("P/E Ratio (Trailing)", fmt_val(info.get("trailingPE"), suffix="x")),
        ("P/E Ratio (Forward)", fmt_val(info.get("forwardPE"), suffix="x")),
        ("EPS (Trailing)", fmt_val(info.get("trailingEps"), prefix="$")),
        ("52-Week High", fmt_val(info.get("fiftyTwoWeekHigh"), prefix="$")),
        ("52-Week Low", fmt_val(info.get("fiftyTwoWeekLow"), prefix="$")),
        ("Revenue (TTM)", fmt_money(info.get("totalRevenue"))),
        ("Gross Margin", fmt_val(info.get("grossMargins", 0) * 100 if info.get("grossMargins") else None, suffix="%", decimals=1)),
        ("Net Profit Margin", fmt_val(info.get("profitMargins", 0) * 100 if info.get("profitMargins") else None, suffix="%", decimals=1)),
        ("Dividend Yield", fmt_val(info.get("dividendYield", 0) * 100 if info.get("dividendYield") else None, suffix="%", decimals=2) if info.get("dividendYield") else "None"),
        ("Beta", fmt_val(info.get("beta"))),
    ]

    for i, (label, value) in enumerate(rows, start=4):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        if i == 4:
            header_style(a)
            header_style(b)
        else:
            a.font = Font(bold=True, size=10)
            b.font = Font(size=10)
            if i % 2 == 0:
                alt_fill(a)
                alt_fill(b)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26


def build_income_sheet(ws, quarterly_income: pd.DataFrame):
    items = {
        "Total Revenue": "Revenue",
        "Gross Profit": "Gross Profit",
        "Operating Income": "Operating Income",
        "Net Income": "Net Income",
    }

    if quarterly_income.empty:
        ws["A1"].value = "No income statement data available."
        return

    cols_chrono = list(quarterly_income.columns[:4][::-1])
    n = len(cols_chrono)

    col_map = []
    c = 2
    for i in range(n):
        if i == 0:
            col_map.append((c, None)); c += 1
        else:
            col_map.append((c, c + 1)); c += 2
    total_cols = c - 1

    title_end = chr(ord("A") + total_cols - 1)
    ws.merge_cells(f"A1:{title_end}1")
    ws["A1"].value = "Quarterly Income Statement  --  Oldest to Newest"
    title_style(ws["A1"], size=13)
    ws.row_dimensions[1].height = 28

    ws.cell(row=3, column=1, value="Item")
    header_style(ws.cell(row=3, column=1))

    for i, qcol in enumerate(cols_chrono):
        val_col, qoq_col = col_map[i]
        label = qcol.strftime("%b %Y") if hasattr(qcol, "strftime") else str(qcol)
        header_style(ws.cell(row=3, column=val_col, value=label))
        if qoq_col:
            c_hdr = ws.cell(row=3, column=qoq_col, value="QoQ %")
            c_hdr.font = Font(bold=True, color=WHITE, size=10)
            c_hdr.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            c_hdr.alignment = Alignment(horizontal="center", vertical="center")

    for row_i, (key, display) in enumerate(items.items(), start=4):
        row_bg = ALT_ROW if row_i % 2 == 0 else WHITE
        label_cell = ws.cell(row=row_i, column=1, value=display)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")

        prev_val = None
        for i, qcol in enumerate(cols_chrono):
            val_col, qoq_col = col_map[i]
            vcell = ws.cell(row=row_i, column=val_col)
            vcell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            cur_val = None
            if key in quarterly_income.index:
                raw = quarterly_income.loc[key, qcol]
                if pd.notna(raw):
                    cur_val = float(raw)
                    vcell.value = cur_val
                    vcell.number_format = '#,##0.00,,"M"'
                    vcell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    vcell.value = "N/A"

            if qoq_col is not None:
                qcell = ws.cell(row=row_i, column=qoq_col)
                qcell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                qcell.alignment = Alignment(horizontal="center", vertical="center")
                if cur_val is not None and prev_val is not None and prev_val != 0:
                    change = (cur_val - prev_val) / abs(prev_val)
                    qcell.value = change
                    qcell.number_format = "0.0%"
                    qcell.font = Font(size=10, bold=True,
                                      color=GREEN if change >= 0 else RED)
                else:
                    qcell.value = "--"
                    qcell.font = Font(size=10, color="888888")

            prev_val = cur_val
        ws.row_dimensions[row_i].height = 18

    ws.column_dimensions["A"].width = 22
    for i in range(n):
        val_col, qoq_col = col_map[i]
        ws.column_dimensions[chr(ord("A") + val_col - 1)].width = 16
        if qoq_col:
            ws.column_dimensions[chr(ord("A") + qoq_col - 1)].width = 10

    note_row = 4 + len(items)
    note = ws.cell(row=note_row, column=1,
                   value="* Values in millions USD  |  QoQ = quarter-over-quarter % change  |  Red = decline")
    note.font = Font(italic=True, color="888888", size=9)


def build_revenue_trend_sheet(ws, quarterly_income: pd.DataFrame):
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Revenue Trend  --  Last 4 Quarters (Oldest -> Newest)"
    title_style(ws["A1"], size=13)
    ws.row_dimensions[1].height = 28

    if "Total Revenue" not in quarterly_income.index:
        ws["A3"].value = "No revenue data available."
        return

    rev = quarterly_income.loc["Total Revenue"].dropna().iloc[:4]
    rev_chrono = rev[::-1]
    rev_list = list(rev_chrono.items())

    headers = ["Quarter", "Revenue (USD)", "QoQ Change", "Trend"]
    for j, h in enumerate(headers, start=1):
        header_style(ws.cell(row=3, column=j, value=h))

    for i, (date, val) in enumerate(rev_list):
        row = i + 4
        label = date.strftime("%b %Y") if hasattr(date, "strftime") else str(date)
        ws.cell(row=row, column=1, value=label)
        money_cell = ws.cell(row=row, column=2, value=float(val))
        money_cell.number_format = "#,##0"
        money_cell.alignment = Alignment(horizontal="right")

        if i > 0:
            prev_val = float(rev_list[i - 1][1])
            change = ((float(val) - prev_val) / abs(prev_val))
            pct_cell = ws.cell(row=row, column=3, value=change)
            pct_cell.number_format = "0.0%"
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.font = Font(color=GREEN if change >= 0 else RED, bold=True)
            trend_cell = ws.cell(row=row, column=4)
            trend_cell.value = "UP" if change >= 0 else "DOWN"
            trend_cell.font = Font(color=GREEN if change >= 0 else RED, bold=True, size=13)
            trend_cell.alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=row, column=3, value="--")
            ws.cell(row=row, column=4, value="--")

        if row % 2 == 0:
            for col_idx in range(1, 5):
                alt_fill(ws.cell(row=row, column=col_idx))

    for col_letter, width in [("A", 14), ("B", 22), ("C", 14), ("D", 10)]:
        ws.column_dimensions[col_letter].width = width


def build_dashboard_sheet(ws, info: dict, quarterly_income: pd.DataFrame,
                          history: pd.DataFrame, commentary: str, news: list = None):
    company = info.get("longName", "N/A")
    symbol = info.get("symbol", "")
    sector = info.get("sector", "")
    exchange = info.get("exchange", "")
    current_price = info.get("currentPrice") or info.get("regularMarketPrice")

    for i, letter in enumerate("ABCDEFGHIJ", start=1):
        ws.column_dimensions[letter].width = 13

    ws.merge_cells("A1:J1")
    h = ws["A1"]
    h.value = f"  {company}  ({symbol})"
    h.font = Font(bold=True, size=16, color=WHITE)
    h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    parts = [p for p in [sector, exchange, f"As of {datetime.now().strftime('%B %d, %Y')}"] if p]
    sub.value = "  " + "  |  ".join(parts)
    sub.font = Font(size=10, color=DARK_BLUE)
    sub.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    yr_change = None
    if not history.empty:
        try:
            yr_change = ((float(history["Close"].iloc[-1]) - float(history["Close"].iloc[0]))
                         / float(history["Close"].iloc[0])) * 100
        except Exception:
            pass

    change_color = GREEN if (yr_change is not None and yr_change >= 0) else RED
    kpis = [
        ("MARKET CAP", fmt_money(info.get("marketCap")), "1F6AA5"),
        ("PRICE", fmt_val(current_price, prefix="$"), "217346"),
        ("P/E RATIO", fmt_val(info.get("trailingPE"), suffix="x"), "5C3D8F"),
        ("GROSS MARGIN", fmt_val(info.get("grossMargins", 0) * 100
                                 if info.get("grossMargins") else None, suffix="%", decimals=1), "0070C0"),
        ("1-YEAR CHANGE", (f"+{yr_change:.1f}%" if yr_change >= 0 else f"{yr_change:.1f}%")
                          if yr_change is not None else "N/A", change_color),
    ]

    kpi_cols = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10)]
    for (label, value, bg), (sc, ec) in zip(kpis, kpi_cols):
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + ec - 1)
        ws.merge_cells(f"{sl}4:{el}4")
        lc = ws.cell(row=4, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{sl}5:{el}5")
        vc = ws.cell(row=5, column=sc, value=value)
        vc.font = Font(bold=True, size=15, color=WHITE)
        vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 10

    # Revenue trend rows 7-12
    ws.merge_cells("A7:J7")
    sec7 = ws["A7"]
    sec7.value = "  QUARTERLY REVENUE TREND"
    sec7.font = Font(bold=True, size=11, color=WHITE)
    sec7.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec7.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22

    for (sl, el, label) in [("A", "B", "Quarter"), ("C", "E", "Revenue (USD)"),
                             ("F", "G", "QoQ Change"), ("H", "J", "Trend Bar")]:
        ws.merge_cells(f"{sl}8:{el}8")
        c = ws[f"{sl}8"]
        c.value = label
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 20

    if "Total Revenue" in quarterly_income.index:
        rev = quarterly_income.loc["Total Revenue"].dropna().iloc[:4]
        rev_chrono = list(rev[::-1].items())
        for i, (date, val) in enumerate(rev_chrono):
            row = 9 + i
            bg = ALT_ROW if i % 2 == 0 else WHITE
            date_label = date.strftime("%b %Y") if hasattr(date, "strftime") else str(date)
            ws.merge_cells(f"A{row}:B{row}")
            dc = ws.cell(row=row, column=1, value=date_label)
            dc.font = Font(size=10, bold=True)
            dc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            dc.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f"C{row}:E{row}")
            rc = ws.cell(row=row, column=3, value=float(val))
            rc.number_format = "#,##0"
            rc.font = Font(size=10)
            rc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            rc.alignment = Alignment(horizontal="right", vertical="center")
            ws.merge_cells(f"F{row}:G{row}")
            qoq = ws.cell(row=row, column=6)
            ws.merge_cells(f"H{row}:J{row}")
            bar = ws.cell(row=row, column=8)
            if i > 0:
                prev_val = float(rev_chrono[i - 1][1])
                change = (float(val) - prev_val) / abs(prev_val)
                qoq.value = change
                qoq.number_format = "0.0%"
                qoq.font = Font(size=10, bold=True, color=GREEN if change >= 0 else RED)
                qoq.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                qoq.alignment = Alignment(horizontal="center", vertical="center")
                bar_len = min(int(abs(change) * 100), 25)
                bar.value = ("+" if change >= 0 else "-") * bar_len
                bar.font = Font(size=9, color=GREEN if change >= 0 else RED, bold=True)
                bar.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                bar.alignment = Alignment(horizontal="left", vertical="center")
            else:
                qoq.value = "Baseline"
                qoq.font = Font(size=10, color="888888", italic=True)
                qoq.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                qoq.alignment = Alignment(horizontal="center", vertical="center")
                bar.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws.row_dimensions[row].height = 20

    ws.row_dimensions[13].height = 10

    # Analyst targets rows 14-18
    ws.merge_cells("A14:J14")
    sec14 = ws["A14"]
    sec14.value = "  ANALYST PRICE TARGETS  (Wall Street Consensus)"
    sec14.font = Font(bold=True, size=11, color=WHITE)
    sec14.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec14.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[14].height = 22

    target_low = info.get("targetLowPrice")
    target_mean = info.get("targetMeanPrice")
    target_high = info.get("targetHighPrice")
    rec_key = (info.get("recommendationKey") or "N/A").upper()
    n_analysts = info.get("numberOfAnalystOpinions", "N/A")

    upside_str = "N/A"
    if target_mean and current_price:
        try:
            upside_pct = ((float(target_mean) - float(current_price)) / float(current_price)) * 100
            direction = "upside" if upside_pct >= 0 else "downside"
            upside_str = f"{'+' if upside_pct >= 0 else ''}{upside_pct:.1f}% implied {direction}"
        except Exception:
            pass

    target_boxes = [
        ("LOW TARGET", f"${target_low:.2f}" if target_low else "N/A", "5C6BC0", 1, 3),
        ("CONSENSUS TARGET", f"${target_mean:.2f}" if target_mean else "N/A", "1F6AA5", 4, 7),
        ("HIGH TARGET", f"${target_high:.2f}" if target_high else "N/A", "2E7D32", 8, 10),
    ]
    for label, value, color, sc, ec in target_boxes:
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + ec - 1)
        ws.merge_cells(f"{sl}15:{el}15")
        lc = ws.cell(row=15, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{sl}16:{el}16")
        vc = ws.cell(row=16, column=sc, value=value)
        vc.font = Font(bold=True, size=16, color=WHITE)
        vc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[15].height = 18
    ws.row_dimensions[16].height = 34

    ws.merge_cells("A17:J17")
    rec_cell = ws["A17"]
    rec_cell.value = f"  Recommendation: {rec_key}  |  {n_analysts} analysts  |  {upside_str}"
    rec_cell.font = Font(size=10, color=DARK_BLUE, italic=True)
    rec_cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    rec_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[17].height = 20
    ws.row_dimensions[18].height = 10

    # Outlook rows 19-22
    ws.merge_cells("A19:J19")
    sec19 = ws["A19"]
    sec19.value = "  OUTLOOK & COMPANY HEALTH"
    sec19.font = Font(bold=True, size=11, color=WHITE)
    sec19.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec19.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[19].height = 22

    ws.merge_cells("A20:J22")
    text_cell = ws["A20"]
    text_cell.value = commentary
    text_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
    text_cell.font = Font(size=10)
    text_cell.fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    for r in range(20, 23):
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[23].height = 8

    # Health indicators rows 24-27
    ws.merge_cells("A24:J24")
    sec24 = ws["A24"]
    sec24.value = "  FINANCIAL HEALTH INDICATORS"
    sec24.font = Font(bold=True, size=11, color=WHITE)
    sec24.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec24.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[24].height = 22

    dte = info.get("debtToEquity")
    roe = info.get("returnOnEquity")
    cr = info.get("currentRatio")
    qr = info.get("quickRatio")
    short_pct = info.get("shortPercentOfFloat")
    short_ratio = info.get("shortRatio")
    ins_pct = info.get("heldPercentInsiders")
    inst_pct = info.get("heldPercentInstitutions")

    def _fv(v, mult=1, prefix="", suffix="", decimals=2):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "N/A"
        return f"{prefix}{float(v) * mult:.{decimals}f}{suffix}"

    health_boxes = [
        ("DEBT / EQUITY", _fv(dte, mult=0.01, suffix="x") if dte else "N/A", "34495E"),
        ("RETURN ON EQUITY", _fv(roe, mult=100, suffix="%", decimals=1), "1A6B3C"),
        ("CURRENT RATIO", _fv(cr, suffix="x"), "1F6AA5"),
        ("QUICK RATIO", _fv(qr, suffix="x"), "0070C0"),
        ("SHORT % FLOAT", _fv(short_pct, mult=100, suffix="%", decimals=1), "7B241C"),
    ]
    for idx, (label, value, bg) in enumerate(health_boxes):
        sc = idx * 2 + 1
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + sc)
        ws.merge_cells(f"{sl}25:{el}25")
        lc = ws.cell(row=25, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"{sl}26:{el}26")
        vc = ws.cell(row=26, column=sc, value=value)
        vc.font = Font(bold=True, size=14, color=WHITE)
        vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[25].height = 18
    ws.row_dimensions[26].height = 30

    ws.merge_cells("A27:J27")
    own = ws["A27"]
    parts_own = []
    if ins_pct is not None:
        parts_own.append(f"Insider ownership: {ins_pct * 100:.1f}%")
    if inst_pct is not None:
        parts_own.append(f"Institutional ownership: {inst_pct * 100:.1f}%")
    if short_ratio is not None:
        parts_own.append(f"Short ratio (days to cover): {short_ratio:.1f}")
    own.value = "  " + "   |   ".join(parts_own) if parts_own else "  Ownership data unavailable."
    own.font = Font(size=10, italic=True, color=DARK_BLUE)
    own.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    own.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[27].height = 20
    ws.row_dimensions[28].height = 8

    # News rows 29+
    news = news or []
    ws.merge_cells("A29:J29")
    sec29 = ws["A29"]
    sec29.value = "  RECENT NEWS"
    sec29.font = Font(bold=True, size=11, color=WHITE)
    sec29.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec29.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[29].height = 22

    if news:
        for i, item in enumerate(news[:6]):
            r = 30 + i
            bg = ALT_ROW if i % 2 == 0 else WHITE
            date_lbl = f"[{item['date']}]" if item.get("date") else ""
            pub_lbl = item.get("publisher", "")
            ws.merge_cells(f"A{r}:B{r}")
            meta = ws[f"A{r}"]
            meta.value = f"{date_lbl}  {pub_lbl}"
            meta.font = Font(size=9, italic=True, color="555555")
            meta.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            meta.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"C{r}:J{r}")
            body = ws[f"C{r}"]
            body.value = item.get("summary", item.get("title", ""))
            body.font = Font(size=10)
            body.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 36
    else:
        ws.merge_cells("A30:J30")
        ws["A30"].value = "  No recent news available."
        ws["A30"].font = Font(size=10, italic=True, color="888888")
        ws.row_dimensions[30].height = 20


def build_industry_sheet(ws, symbol: str, info: dict, peers: list):
    if not peers:
        ws["A1"].value = "Industry peer data not available."
        return

    industry_label = info.get("industry", info.get("industryKey", "Industry"))
    company_name = info.get("longName", symbol)
    n_peers = len(peers)

    def pct(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "N/A"
        return f"{val * 100:.1f}%"

    def pe(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "N/A"
        return f"{float(val):.1f}x"

    def diff_str(company_val, avg_val):
        if company_val is None or avg_val is None or avg_val == 0:
            return "N/A"
        d = (company_val - avg_val) / abs(avg_val) * 100
        return f"{'+' if d >= 0 else ''}{d:.1f}%"

    total_mktcap = sum(p["marketCap"] for p in peers if p["marketCap"])
    avg_trailing_pe = _safe_avg([p["trailingPE"] for p in peers])
    avg_forward_pe = _safe_avg([p["forwardPE"] for p in peers])
    avg_gross_margin = _safe_avg([p["grossMargins"] for p in peers])
    avg_net_margin = _safe_avg([p["profitMargins"] for p in peers])
    avg_rev_growth = _safe_avg([p["revenueGrowth"] for p in peers])
    avg_52wk_chg = _safe_avg([p["fiftyTwoWeekChange"] for p in peers])

    target = next((p for p in peers if p["is_target"]), peers[0])

    col_widths = {"A": 30, "B": 16, "C": 16, "D": 16, "E": 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    ws.merge_cells(f"A{row}:E{row}")
    h = ws[f"A{row}"]
    h.value = f"  {industry_label.upper()}  --  Industry Benchmarks"
    h.font = Font(bold=True, size=14, color=WHITE)
    h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    sub = ws[f"A{row}"]
    sub.value = f"  {n_peers} companies analysed  |  Source: Yahoo Finance  |  {datetime.now().strftime('%B %d, %Y')}"
    sub.font = Font(size=9, italic=True, color=DARK_BLUE)
    sub.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 2

    ws.merge_cells(f"A{row}:E{row}")
    sec2 = ws[f"A{row}"]
    sec2.value = f"  {company_name}  vs.  Industry Average"
    sec2.font = Font(bold=True, size=11, color=WHITE)
    sec2.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, label in [(1, "Metric"), (2, company_name[:20]), (3, "Industry Avg"),
                       (4, "vs. Avg"), (5, "Signal")]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    metrics = [
        ("Market Cap", target["marketCap"], total_mktcap / n_peers, fmt_money, True),
        ("P/E Ratio (Trailing)", target["trailingPE"], avg_trailing_pe, pe, False),
        ("P/E Ratio (Forward)", target["forwardPE"], avg_forward_pe, pe, False),
        ("Gross Margin", target["grossMargins"], avg_gross_margin, pct, True),
        ("Net Profit Margin", target["profitMargins"], avg_net_margin, pct, True),
        ("Revenue Growth (YoY)", target["revenueGrowth"], avg_rev_growth, pct, True),
        ("52-Week Price Change", target["fiftyTwoWeekChange"], avg_52wk_chg, pct, True),
    ]

    for i, (label, co_val, avg_val, fmt_fn, higher_better) in enumerate(metrics):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        fill_c, font_c = _cmp_color(co_val, avg_val, higher_better)

        def write(col, value, bold=False, fill=bg, color="000000", align="left"):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(size=10, bold=bold, color=color)
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center")

        write(1, label, bold=True)
        write(2, fmt_fn(co_val), align="center")
        write(3, fmt_fn(avg_val), align="center")
        write(4, diff_str(co_val, avg_val), fill=fill_c, color=font_c, align="center")

        if fill_c == CELL_GREEN:
            sig, sig_color = "Above avg", TEXT_GREEN
        elif fill_c == CELL_RED:
            sig, sig_color = "Below avg", TEXT_RED
        else:
            sig, sig_color = "At par", "666666"
        c5 = ws.cell(row=row, column=5, value=sig)
        c5.font = Font(size=10, bold=True, color=sig_color)
        c5.fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type="solid")
        c5.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    peer_cols = {"A": 32, "B": 10, "C": 14, "D": 10, "E": 10,
                 "F": 14, "G": 14, "H": 14, "I": 14}
    for col, w in peer_cols.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells(f"A{row}:I{row}")
    sec3 = ws[f"A{row}"]
    sec3.value = "  Full Peer Comparison"
    sec3.font = Font(bold=True, size=11, color=WHITE)
    sec3.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec3.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    peer_headers = ["Company", "Ticker", "Market Cap", "P/E Trail",
                    "P/E Fwd", "Gross Margin", "Net Margin", "Rev Growth", "52-Wk Chg"]
    for j, h_label in enumerate(peer_headers, start=1):
        c = ws.cell(row=row, column=j, value=h_label)
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center" if j > 1 else "left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    peers_sorted = sorted(peers, key=lambda p: p["marketCap"] or 0, reverse=True)

    metric_keys = ["marketCap", "trailingPE", "forwardPE",
                   "grossMargins", "profitMargins", "revenueGrowth", "fiftyTwoWeekChange"]
    metric_col = {k: j for j, k in enumerate(metric_keys, start=3)}

    def _clean(vals):
        return [v for v in vals if v is not None and not (isinstance(v, float) and pd.isna(v))]

    col_max = {k: max(_clean([p[k] for p in peers_sorted]), default=None) for k in metric_keys}
    col_min = {k: min(_clean([p[k] for p in peers_sorted]), default=None) for k in metric_keys}

    display_fns = {
        "marketCap": fmt_money,
        "trailingPE": pe,
        "forwardPE": pe,
        "grossMargins": pct,
        "profitMargins": pct,
        "revenueGrowth": pct,
        "fiftyTwoWeekChange": pct,
    }

    for i, p in enumerate(peers_sorted):
        is_tgt = p["is_target"]
        row_bg = "D6E4F0" if is_tgt else (ALT_ROW if i % 2 == 0 else WHITE)

        for col, val, align in [(1, p["name"][:40], "left"), (2, p["symbol"], "center")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=10, bold=is_tgt, color=DARK_BLUE if is_tgt else "000000")
            c.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center")

        for key in metric_keys:
            col = metric_col[key]
            raw = p[key]
            display = display_fns[key](raw)
            is_max = (raw is not None and col_max[key] is not None and
                      not (isinstance(raw, float) and pd.isna(raw)) and raw == col_max[key])
            is_min = (raw is not None and col_min[key] is not None and
                      not (isinstance(raw, float) and pd.isna(raw)) and raw == col_min[key])

            if is_max:
                cell_bg, cell_fc = CELL_GREEN, TEXT_GREEN
            elif is_min:
                cell_bg, cell_fc = CELL_RED, TEXT_RED
            else:
                cell_bg, cell_fc = row_bg, (DARK_BLUE if is_tgt else "000000")

            c = ws.cell(row=row, column=col, value=display)
            c.font = Font(size=10, bold=(is_tgt or is_max or is_min), color=cell_fc)
            c.fill = PatternFill(start_color=cell_bg, end_color=cell_bg, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    ws.merge_cells(f"A{row}:I{row}")
    note = ws[f"A{row}"]
    note.value = ("  * Target company highlighted in blue.  "
                  "Green cell = highest in column.  Red cell = lowest in column.  "
                  "All data from Yahoo Finance.")
    note.font = Font(size=8, italic=True, color="888888")
    ws.row_dimensions[row].height = 14


def build_full_workbook(symbol: str, info: dict, quarterly_income, history,
                        commentary: str, news: list, peers: list) -> io.BytesIO:
    """Build the complete workbook and return it as an in-memory BytesIO buffer."""
    wb = Workbook()

    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    build_dashboard_sheet(ws_dash, info, quarterly_income, history, commentary, news=news)

    ws_industry = wb.create_sheet("Industry Comparison")
    build_industry_sheet(ws_industry, symbol, info, peers)

    ws_overview = wb.create_sheet("Overview")
    build_overview_sheet(ws_overview, info)

    ws_income = wb.create_sheet("Income Statement")
    build_income_sheet(ws_income, quarterly_income)

    ws_trend = wb.create_sheet("Revenue Trend")
    build_revenue_trend_sheet(ws_trend, quarterly_income)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

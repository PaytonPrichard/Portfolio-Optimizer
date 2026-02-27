#!/usr/bin/env python3
"""
MarketMosaic — CLI financial analyzer
Usage:
    python company_financials.py AAPL
    python company_financials.py          # will prompt for ticker

Output: outputs/<TICKER>_financials.xlsx
"""

import os
import sys
from datetime import datetime

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_DIR = "outputs"

# ── Colours ────────────────────────────────────────────────────────────────────
DARK_BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "F0F4F8"
WHITE = "FFFFFF"
GREEN = "00703C"
RED = "CC0000"


# ── Helpers ────────────────────────────────────────────────────────────────────

def fmt_money(value) -> str:
    """Format a dollar value into readable billions / millions."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "N/A"
    try:
        value = float(value)
    except (TypeError, ValueError):
        return "N/A"
    if abs(value) >= 1e12:
        return f"${value / 1e12:.2f}T"
    if abs(value) >= 1e9:
        return f"${value / 1e9:.2f}B"
    if abs(value) >= 1e6:
        return f"${value / 1e6:.2f}M"
    return f"${value:,.0f}"


def fmt_val(value, prefix="", suffix="", decimals=2) -> str:
    """Generic value formatter."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "N/A"
    try:
        return f"{prefix}{float(value):.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return "N/A"


def header_style(cell, bg=DARK_BLUE, fg=WHITE):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def alt_fill(cell):
    cell.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")


def title_style(cell, size=14):
    cell.font = Font(bold=True, size=size, color=DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── Data fetching ───────────────────────────────────────────────────────────────

def fetch_data(symbol: str):
    """Return (ticker, info dict, quarterly income DataFrame, 1-yr price history)."""
    print(f"\nFetching data for {symbol}...")
    ticker = yf.Ticker(symbol)
    info = ticker.info or {}
    quarterly_income = ticker.quarterly_income_stmt   # columns = quarters (newest first)
    history = ticker.history(period="1y")
    return ticker, info, quarterly_income, history


def fetch_recent_news(symbol: str, n: int = 8) -> list:
    """Return up to n recent news dicts for the ticker (title, publisher, date, link)."""
    try:
        raw = yf.Ticker(symbol).news or []
        out = []
        for item in raw[:n]:
            title = item.get("title") or item.get("content", {}).get("title", "")
            publisher = item.get("publisher") or item.get("content", {}).get("provider", {}).get("displayName", "")
            ts = item.get("providerPublishTime") or item.get("content", {}).get("pubDate")
            if ts:
                try:
                    date_str = datetime.fromtimestamp(int(ts)).strftime("%b %d, %Y")
                except Exception:
                    date_str = str(ts)[:10]
            else:
                date_str = ""
            if title:
                out.append({"title": title, "publisher": publisher, "date": date_str})
        return out
    except Exception:
        return []


# ── Plain-English summary ───────────────────────────────────────────────────────

def generate_summary(info: dict, quarterly_income: pd.DataFrame, history: pd.DataFrame) -> str:
    """Build a plain-English paragraph summarising financial trends."""
    company = info.get("longName", "This company")
    sentences = []

    # Revenue trend
    if "Total Revenue" in quarterly_income.index:
        rev = quarterly_income.loc["Total Revenue"].dropna()
        if len(rev) >= 2:
            latest, prev = float(rev.iloc[0]), float(rev.iloc[1])
            pct = ((latest - prev) / abs(prev)) * 100
            direction = "grew" if pct >= 0 else "declined"
            sentences.append(
                f"{company} revenue {direction} {abs(pct):.1f}% quarter-over-quarter "
                f"({fmt_money(prev)} -> {fmt_money(latest)})."
            )
        if len(rev) >= 4:
            oldest, latest_4 = float(rev.iloc[3]), float(rev.iloc[0])
            overall = ((latest_4 - oldest) / abs(oldest)) * 100
            trend = "upward" if overall >= 0 else "downward"
            sentences.append(
                f"The 4-quarter revenue trend is {trend} "
                f"({fmt_money(oldest)} -> {fmt_money(latest_4)}, "
                f"{abs(overall):.1f}% {'growth' if overall >= 0 else 'decline'} overall)."
            )

    # Gross margin
    gm = info.get("grossMargins")
    if gm:
        sentences.append(f"Gross margin stands at {gm * 100:.1f}%.")

    # Net income
    if "Net Income" in quarterly_income.index:
        ni = quarterly_income.loc["Net Income"].dropna()
        if len(ni) >= 1:
            latest_ni = float(ni.iloc[0])
            if latest_ni > 0:
                sentences.append(
                    f"The most recent quarter shows positive net income of {fmt_money(latest_ni)}."
                )
            else:
                sentences.append(
                    f"The most recent quarter shows a net loss of {fmt_money(abs(latest_ni))}, "
                    f"indicating the company is not yet profitable."
                )

    # Stock price
    if not history.empty:
        start = float(history["Close"].iloc[0])
        end = float(history["Close"].iloc[-1])
        pct = ((end - start) / start) * 100
        direction = "gained" if pct >= 0 else "lost"
        sentences.append(
            f"Over the past year the stock has {direction} {abs(pct):.1f}% "
            f"(${start:.2f} -> ${end:.2f})."
        )

    # P/E context
    pe = info.get("trailingPE")
    if pe and not pd.isna(pe):
        pe = float(pe)
        if pe < 15:
            comment = "suggesting the stock may be undervalued relative to the broader market"
        elif pe < 30:
            comment = "in line with typical market valuations"
        else:
            comment = "reflecting high growth expectations from investors"
        sentences.append(f"The trailing P/E ratio is {pe:.1f}x, {comment}.")

    if not sentences:
        return "Insufficient data available to generate a trend summary."
    return " ".join(sentences)


def generate_ai_commentary(info: dict, quarterly_income: pd.DataFrame,
                           history: pd.DataFrame, news: list = None) -> str:
    """
    Generate analyst commentary using Claude Haiku (5 sentences: financials + news context).
    Falls back to generate_summary() if ANTHROPIC_API_KEY is not set or unavailable.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("  (ANTHROPIC_API_KEY not set — using rule-based commentary)")
        return generate_summary(info, quarterly_income, history)

    try:
        import anthropic
    except ImportError:
        print("  (anthropic package not installed — using rule-based commentary)")
        return generate_summary(info, quarterly_income, history)

    news = news or []

    company = info.get("longName", "This company")
    symbol = info.get("symbol", "")
    current_price = info.get("currentPrice") or info.get("regularMarketPrice")
    target_mean = info.get("targetMeanPrice")
    target_low = info.get("targetLowPrice")
    target_high = info.get("targetHighPrice")
    rec = (info.get("recommendationKey") or "N/A").upper()
    n_analysts = info.get("numberOfAnalystOpinions", "N/A")

    rev_lines = []
    if "Total Revenue" in quarterly_income.index:
        rev = quarterly_income.loc["Total Revenue"].dropna().iloc[:4]
        for date, val in rev.items():
            label = date.strftime("%b %Y") if hasattr(date, "strftime") else str(date)
            rev_lines.append(f"  {label}: {fmt_money(float(val))}")

    ni_line = ""
    if "Net Income" in quarterly_income.index:
        ni = quarterly_income.loc["Net Income"].dropna()
        if len(ni) >= 1:
            ni_line = f"Net income (latest quarter): {fmt_money(float(ni.iloc[0]))}"

    price_change_line = ""
    if not history.empty:
        try:
            start = float(history["Close"].iloc[0])
            end = float(history["Close"].iloc[-1])
            pct = ((end - start) / start) * 100
            price_change_line = f"1-year stock change: {'+' if pct >= 0 else ''}{pct:.1f}%"
        except Exception:
            pass

    upside_line = ""
    if target_mean and current_price:
        try:
            upside_pct = ((float(target_mean) - float(current_price)) / float(current_price)) * 100
            upside_line = f"Implied upside to consensus: {'+' if upside_pct >= 0 else ''}{upside_pct:.1f}%"
        except Exception:
            pass

    news_lines = [
        f"  - {n['date']}  {n['publisher']}:  {n['title']}"
        for n in news if n.get("title")
    ]

    data_block = "\n".join(filter(None, [
        f"Company: {company} ({symbol})",
        f"Current price: ${current_price}",
        f"52-week range: ${info.get('fiftyTwoWeekLow', 'N/A')} - ${info.get('fiftyTwoWeekHigh', 'N/A')}",
        price_change_line,
        f"P/E ratio (trailing): {fmt_val(info.get('trailingPE'), suffix='x')}",
        f"Gross margin: {fmt_val(info.get('grossMargins', 0) * 100 if info.get('grossMargins') else None, suffix='%', decimals=1)}",
        f"Net profit margin: {fmt_val(info.get('profitMargins', 0) * 100 if info.get('profitMargins') else None, suffix='%', decimals=1)}",
        ni_line,
        "Quarterly revenue (newest first):",
        *rev_lines,
        f"Analyst consensus target: ${target_mean} (low: ${target_low}, high: ${target_high})",
        f"Analyst recommendation: {rec} ({n_analysts} analysts)",
        upside_line,
        ("Recent news headlines:\n" + "\n".join(news_lines)) if news_lines else "",
    ]))

    prompt = (
        "You are a concise equity research analyst writing a company assessment "
        "for a financial dashboard.\n\n"
        "Based on the data and recent news below, write exactly 5 sentences with NO headers or labels:\n"
        "1. Company health: summarize operational performance (revenue trend, margins, profitability).\n"
        "2. Price outlook: reference the analyst consensus target and implied upside/downside.\n"
        "3. Key financial risk or opportunity evident in the numbers.\n"
        "4. Recent company-specific development: draw on the news headlines (e.g. M&A, leadership, "
        "product launch, regulatory event). If no relevant news, note the most significant recent story.\n"
        "5. Broader industry or macro context relevant to this company right now.\n\n"
        "Rules: be specific and reference actual figures or news items. Keep each sentence under 50 words. "
        "No bullet points, no numbered list, no section labels. Just 5 plain sentences in a paragraph.\n\n"
        f"Financial data and news:\n{data_block}"
    )

    try:
        print("  Generating AI commentary...")
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        return response.content[0].text.strip()
    except Exception as exc:
        print(f"  AI commentary failed ({exc}) — falling back to rule-based summary.")
        return generate_summary(info, quarterly_income, history)


def generate_news_summaries(news: list, company_name: str) -> list:
    """
    Add a 'summary' field to each news item using a single Claude Haiku call.
    Falls back to using the headline as the summary if no API key is set.
    """
    if not news:
        return news

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        for item in news:
            item["summary"] = item.get("title", "")
        return news

    try:
        import anthropic, re
    except ImportError:
        for item in news:
            item["summary"] = item.get("title", "")
        return news

    headlines_block = "\n".join(
        f"{i + 1}. [{item.get('publisher', '')}]  {item['title']}"
        for i, item in enumerate(news)
        if item.get("title")
    )
    n = len(news)
    prompt = (
        f"These are recent news headlines related to {company_name}.\n"
        f"For each headline write ONE sentence (max 30 words) summarising what the article is about.\n\n"
        f"{headlines_block}\n\n"
        f"Return exactly {n} lines numbered 1-{n}. One sentence per line. No extra text."
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=400,
            messages=[{"role": "user", "content": prompt}],
        )
        lines = [l.strip() for l in resp.content[0].text.strip().split("\n") if l.strip()]
        summaries = [re.sub(r"^\d+[\.\)]\s*", "", l) for l in lines]
        for i, item in enumerate(news):
            item["summary"] = summaries[i] if i < len(summaries) else item.get("title", "")
    except Exception:
        for item in news:
            item["summary"] = item.get("title", "")

    return news


def _news_headlines_text(news: list) -> str:
    """Format news list as plain-text headlines for rule-based fallback."""
    if not news:
        return ""
    lines = ["Recent Headlines:"]
    for n in news[:5]:
        date = f"[{n['date']}]  " if n.get("date") else ""
        pub  = f"{n['publisher']}:  " if n.get("publisher") else ""
        lines.append(f"  - {date}{pub}{n['title']}")
    return "\n".join(lines)


# ── Excel builder ───────────────────────────────────────────────────────────────

def build_overview_sheet(ws, info: dict):
    ws.merge_cells("A1:C1")
    title_style(ws["A1"])
    company = info.get("longName", "N/A")
    symbol = info.get("symbol", "")
    ws["A1"].value = f"{company} ({symbol})  —  Financial Overview"
    ws.row_dimensions[1].height = 32

    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=9)
    ws.row_dimensions[2].height = 14

    rows = [
        ("Metric", "Value"),
        ("Company", company),
        ("Sector", info.get("sector", "N/A")),
        ("Industry", info.get("industry", "N/A")),
        ("Market Cap", fmt_money(info.get("marketCap"))),
        ("Current Price", fmt_val(info.get("currentPrice") or info.get("regularMarketPrice"), prefix="$")),
        ("P/E Ratio (Trailing)", fmt_val(info.get("trailingPE"), suffix="x")),
        ("P/E Ratio (Forward)", fmt_val(info.get("forwardPE"), suffix="x")),
        ("EPS (Trailing)", fmt_val(info.get("trailingEps"), prefix="$")),
        ("52-Week High", fmt_val(info.get("fiftyTwoWeekHigh"), prefix="$")),
        ("52-Week Low", fmt_val(info.get("fiftyTwoWeekLow"), prefix="$")),
        ("Revenue (TTM)", fmt_money(info.get("totalRevenue"))),
        ("Gross Margin", fmt_val(info.get("grossMargins", 0) * 100 if info.get("grossMargins") else None, suffix="%", decimals=1)),
        ("Net Profit Margin", fmt_val(info.get("profitMargins", 0) * 100 if info.get("profitMargins") else None, suffix="%", decimals=1)),
        ("Dividend Yield", fmt_val(info.get("dividendYield", 0) * 100 if info.get("dividendYield") else None, suffix="%", decimals=2) if info.get("dividendYield") else "None"),
        ("Beta", fmt_val(info.get("beta"))),
    ]

    for i, (label, value) in enumerate(rows, start=4):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=value)
        if i == 4:
            header_style(a)
            header_style(b)
        else:
            a.font = Font(bold=True, size=10)
            b.font = Font(size=10)
            if i % 2 == 0:
                alt_fill(a)
                alt_fill(b)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 26


def build_income_sheet(ws, quarterly_income: pd.DataFrame):
    """
    Chronological order (oldest → newest, left → right).
    Interleaved QoQ % change columns between each quarter; red font if negative.
    """
    items = {
        "Total Revenue":    "Revenue",
        "Gross Profit":     "Gross Profit",
        "Operating Income": "Operating Income",
        "Net Income":       "Net Income",
    }

    if quarterly_income.empty:
        ws["A1"].value = "No income statement data available."
        return

    # Chronological: oldest quarter first
    cols_chrono = list(quarterly_income.columns[:4][::-1])
    n = len(cols_chrono)

    # Build column index map:
    # Q0 → val_col=2 (no QoQ)
    # Q1 → val_col=3, qoq_col=4
    # Q2 → val_col=5, qoq_col=6
    # Q3 → val_col=7, qoq_col=8
    col_map = []          # list of (val_col, qoq_col_or_None)
    c = 2
    for i in range(n):
        if i == 0:
            col_map.append((c, None)); c += 1
        else:
            col_map.append((c, c + 1)); c += 2
    total_cols = c - 1    # last used column index

    # Title
    title_end = chr(ord("A") + total_cols - 1)
    ws.merge_cells(f"A1:{title_end}1")
    ws["A1"].value = "Quarterly Income Statement  —  Oldest to Newest"
    title_style(ws["A1"], size=13)
    ws.row_dimensions[1].height = 28

    # Header row (row 3)
    ws.cell(row=3, column=1, value="Item")
    header_style(ws.cell(row=3, column=1))

    for i, qcol in enumerate(cols_chrono):
        val_col, qoq_col = col_map[i]
        label = qcol.strftime("%b %Y") if hasattr(qcol, "strftime") else str(qcol)
        header_style(ws.cell(row=3, column=val_col, value=label))
        if qoq_col:
            c_hdr = ws.cell(row=3, column=qoq_col, value="QoQ %")
            c_hdr.font  = Font(bold=True, color=WHITE, size=10)
            c_hdr.fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            c_hdr.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row_i, (key, display) in enumerate(items.items(), start=4):
        row_bg = ALT_ROW if row_i % 2 == 0 else WHITE

        label_cell = ws.cell(row=row_i, column=1, value=display)
        label_cell.font = Font(bold=True, size=10)
        label_cell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")

        prev_val = None
        for i, qcol in enumerate(cols_chrono):
            val_col, qoq_col = col_map[i]

            # Value cell
            vcell = ws.cell(row=row_i, column=val_col)
            vcell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            cur_val = None
            if key in quarterly_income.index:
                raw = quarterly_income.loc[key, qcol]
                if pd.notna(raw):
                    cur_val = float(raw)
                    vcell.value = cur_val
                    vcell.number_format = '#,##0.00,,"M"'
                    vcell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    vcell.value = "N/A"

            # QoQ % cell
            if qoq_col is not None:
                qcell = ws.cell(row=row_i, column=qoq_col)
                qcell.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
                qcell.alignment = Alignment(horizontal="center", vertical="center")
                if cur_val is not None and prev_val is not None and prev_val != 0:
                    change = (cur_val - prev_val) / abs(prev_val)
                    qcell.value = change
                    qcell.number_format = "0.0%"
                    qcell.font = Font(size=10, bold=True,
                                      color=GREEN if change >= 0 else RED)
                else:
                    qcell.value = "—"
                    qcell.font = Font(size=10, color="888888")

            prev_val = cur_val
        ws.row_dimensions[row_i].height = 18

    # Column widths
    ws.column_dimensions["A"].width = 22
    for i in range(n):
        val_col, qoq_col = col_map[i]
        ws.column_dimensions[chr(ord("A") + val_col - 1)].width = 16
        if qoq_col:
            ws.column_dimensions[chr(ord("A") + qoq_col - 1)].width = 10

    note_row = 4 + len(items)
    note = ws.cell(row=note_row, column=1,
                   value="* Values in millions USD  |  QoQ = quarter-over-quarter % change  |  Red = decline")
    note.font = Font(italic=True, color="888888", size=9)


def build_revenue_trend_sheet(ws, quarterly_income: pd.DataFrame):
    ws.merge_cells("A1:D1")
    ws["A1"].value = "Revenue Trend  —  Last 4 Quarters (Oldest → Newest)"
    title_style(ws["A1"], size=13)
    ws.row_dimensions[1].height = 28

    if "Total Revenue" not in quarterly_income.index:
        ws["A3"].value = "No revenue data available."
        return

    rev = quarterly_income.loc["Total Revenue"].dropna().iloc[:4]
    # Reverse so oldest quarter is first (chronological order)
    rev_chrono = rev[::-1]
    rev_list = list(rev_chrono.items())

    headers = ["Quarter", "Revenue (USD)", "QoQ Change", "Trend"]
    for j, h in enumerate(headers, start=1):
        header_style(ws.cell(row=3, column=j, value=h))

    for i, (date, val) in enumerate(rev_list):
        row = i + 4
        label = date.strftime("%b %Y") if hasattr(date, "strftime") else str(date)
        ws.cell(row=row, column=1, value=label)

        money_cell = ws.cell(row=row, column=2, value=float(val))
        money_cell.number_format = "#,##0"
        money_cell.alignment = Alignment(horizontal="right")

        if i > 0:
            prev_val = float(rev_list[i - 1][1])
            change = ((float(val) - prev_val) / abs(prev_val))
            pct_cell = ws.cell(row=row, column=3, value=change)
            pct_cell.number_format = "0.0%"
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.font = Font(color=GREEN if change >= 0 else RED, bold=True)

            trend_cell = ws.cell(row=row, column=4)
            trend_cell.value = "▲" if change >= 0 else "▼"
            trend_cell.font = Font(color=GREEN if change >= 0 else RED, bold=True, size=13)
            trend_cell.alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=row, column=3, value="—")
            ws.cell(row=row, column=4, value="—")

        if row % 2 == 0:
            for col_idx in range(1, 5):
                alt_fill(ws.cell(row=row, column=col_idx))

    for col_letter, width in [("A", 14), ("B", 22), ("C", 14), ("D", 10)]:
        ws.column_dimensions[col_letter].width = width


def build_summary_sheet(ws, company_name: str, summary: str, news: list = None):
    for col_letter in ["A", "B", "C", "D"]:
        ws.column_dimensions[col_letter].width = 24

    ws.merge_cells("A1:D1")
    ws["A1"].value = f"Trend Summary  —  {company_name}"
    title_style(ws["A1"], size=13)
    ws.row_dimensions[1].height = 28

    ws["A3"].value = "Analysis & Outlook"
    ws["A3"].font = Font(bold=True, size=11, color=DARK_BLUE)
    ws["A3"].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ws.row_dimensions[3].height = 18

    ws.merge_cells("A4:D12")
    cell = ws["A4"]
    cell.value = summary
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.font = Font(size=11)
    ws.row_dimensions[4].height = 120

    # Recent Headlines section (shown when AI commentary not used, but always useful)
    news = news or []
    if news:
        ws["A14"].value = "Recent Headlines"
        ws["A14"].font = Font(bold=True, size=11, color=DARK_BLUE)
        ws["A14"].fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        ws.row_dimensions[14].height = 18

        for i, n in enumerate(news[:6], start=15):
            date_pub = f"[{n['date']}]  {n['publisher']}  —  " if n.get("date") else ""
            ws.merge_cells(f"A{i}:D{i}")
            c = ws[f"A{i}"]
            c.value = date_pub + n["title"]
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            bg = ALT_ROW if i % 2 == 0 else WHITE
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            ws.row_dimensions[i].height = 28


def build_dashboard_sheet(ws, info: dict, quarterly_income: pd.DataFrame,
                          history: pd.DataFrame, commentary: str, news: list = None):
    """Single-page visual dashboard: KPI boxes, revenue trend, analyst targets, outlook."""

    company = info.get("longName", "N/A")
    symbol = info.get("symbol", "")
    sector = info.get("sector", "")
    exchange = info.get("exchange", "")
    current_price = info.get("currentPrice") or info.get("regularMarketPrice")

    # Column widths — 10 cols (A–J), each 13 wide ≈ one landscape page
    for i, letter in enumerate("ABCDEFGHIJ", start=1):
        ws.column_dimensions[letter].width = 13

    # ── Header ───────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    h = ws["A1"]
    h.value = f"  {company}  ({symbol})"
    h.font = Font(bold=True, size=16, color=WHITE)
    h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    sub = ws["A2"]
    parts = [p for p in [sector, exchange, f"As of {datetime.now().strftime('%B %d, %Y')}"] if p]
    sub.value = "  " + "  |  ".join(parts)
    sub.font = Font(size=10, color=DARK_BLUE)
    sub.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8  # spacer

    # ── KPI Boxes (rows 4–5) ─────────────────────────────────────────────────────
    # 5 boxes × 2 cols each = 10 cols (A:B, C:D, E:F, G:H, I:J)
    yr_change = None
    if not history.empty:
        try:
            yr_change = ((float(history["Close"].iloc[-1]) - float(history["Close"].iloc[0]))
                         / float(history["Close"].iloc[0])) * 100
        except Exception:
            pass

    change_color = GREEN if (yr_change is not None and yr_change >= 0) else RED
    kpis = [
        ("MARKET CAP",    fmt_money(info.get("marketCap")),                                          "1F6AA5"),
        ("PRICE",         fmt_val(current_price, prefix="$"),                                        "217346"),
        ("P/E RATIO",     fmt_val(info.get("trailingPE"), suffix="x"),                               "5C3D8F"),
        ("GROSS MARGIN",  fmt_val(info.get("grossMargins", 0) * 100
                                  if info.get("grossMargins") else None, suffix="%", decimals=1),    "0070C0"),
        ("1-YEAR CHANGE", (f"+{yr_change:.1f}%" if yr_change >= 0 else f"{yr_change:.1f}%")
                          if yr_change is not None else "N/A",                                       change_color),
    ]

    kpi_cols = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10)]  # (start_col, end_col) 1-indexed
    for (label, value, bg), (sc, ec) in zip(kpis, kpi_cols):
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + ec - 1)
        ws.merge_cells(f"{sl}4:{el}4")
        lc = ws.cell(row=4, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{sl}5:{el}5")
        vc = ws.cell(row=5, column=sc, value=value)
        vc.font = Font(bold=True, size=15, color=WHITE)
        vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 34
    ws.row_dimensions[6].height = 10  # spacer

    # ── Revenue Trend (rows 7–13) ────────────────────────────────────────────────
    ws.merge_cells("A7:J7")
    sec7 = ws["A7"]
    sec7.value = "  QUARTERLY REVENUE TREND"
    sec7.font = Font(bold=True, size=11, color=WHITE)
    sec7.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec7.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[7].height = 22

    # Sub-headers
    for (sl, el, label) in [("A", "B", "Quarter"), ("C", "E", "Revenue (USD)"),
                             ("F", "G", "QoQ Change"), ("H", "J", "Trend Bar")]:
        ws.merge_cells(f"{sl}8:{el}8")
        c = ws[f"{sl}8"]
        c.value = label
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 20

    if "Total Revenue" in quarterly_income.index:
        rev = quarterly_income.loc["Total Revenue"].dropna().iloc[:4]
        rev_chrono = list(rev[::-1].items())  # oldest → newest
        for i, (date, val) in enumerate(rev_chrono):
            row = 9 + i
            bg = ALT_ROW if i % 2 == 0 else WHITE
            date_label = date.strftime("%b %Y") if hasattr(date, "strftime") else str(date)

            ws.merge_cells(f"A{row}:B{row}")
            dc = ws.cell(row=row, column=1, value=date_label)
            dc.font = Font(size=10, bold=True)
            dc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            dc.alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells(f"C{row}:E{row}")
            rc = ws.cell(row=row, column=3, value=float(val))
            rc.number_format = "#,##0"
            rc.font = Font(size=10)
            rc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            rc.alignment = Alignment(horizontal="right", vertical="center")

            ws.merge_cells(f"F{row}:G{row}")
            qoq = ws.cell(row=row, column=6)
            ws.merge_cells(f"H{row}:J{row}")
            bar = ws.cell(row=row, column=8)

            if i > 0:
                prev_val = float(rev_chrono[i - 1][1])
                change = (float(val) - prev_val) / abs(prev_val)
                qoq.value = change
                qoq.number_format = "0.0%"
                qoq.font = Font(size=10, bold=True, color=GREEN if change >= 0 else RED)
                qoq.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                qoq.alignment = Alignment(horizontal="center", vertical="center")
                bar_len = min(int(abs(change) * 100), 25)
                bar.value = ("+" if change >= 0 else "-") * bar_len
                bar.font = Font(size=9, color=GREEN if change >= 0 else RED, bold=True)
                bar.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                bar.alignment = Alignment(horizontal="left", vertical="center")
            else:
                qoq.value = "Baseline"
                qoq.font = Font(size=10, color="888888", italic=True)
                qoq.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                qoq.alignment = Alignment(horizontal="center", vertical="center")
                bar.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

            ws.row_dimensions[row].height = 20

    ws.row_dimensions[13].height = 10  # spacer

    # ── Analyst Price Targets (rows 14–18) ───────────────────────────────────────
    ws.merge_cells("A14:J14")
    sec14 = ws["A14"]
    sec14.value = "  ANALYST PRICE TARGETS  (Wall Street Consensus)"
    sec14.font = Font(bold=True, size=11, color=WHITE)
    sec14.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec14.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[14].height = 22

    target_low = info.get("targetLowPrice")
    target_mean = info.get("targetMeanPrice")
    target_high = info.get("targetHighPrice")
    rec_key = (info.get("recommendationKey") or "N/A").upper()
    n_analysts = info.get("numberOfAnalystOpinions", "N/A")

    upside_str = "N/A"
    if target_mean and current_price:
        try:
            upside_pct = ((float(target_mean) - float(current_price)) / float(current_price)) * 100
            direction = "upside" if upside_pct >= 0 else "downside"
            upside_str = f"{'+' if upside_pct >= 0 else ''}{upside_pct:.1f}% implied {direction}"
        except Exception:
            pass

    # 3 target boxes: A:C  D:G  H:J  (3 + 4 + 3 = 10 cols)
    target_boxes = [
        ("LOW TARGET",       f"${target_low:.2f}" if target_low else "N/A",  "5C6BC0", 1,  3),
        ("CONSENSUS TARGET", f"${target_mean:.2f}" if target_mean else "N/A", "1F6AA5", 4,  7),
        ("HIGH TARGET",      f"${target_high:.2f}" if target_high else "N/A", "2E7D32", 8, 10),
    ]
    for label, value, color, sc, ec in target_boxes:
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + ec - 1)
        ws.merge_cells(f"{sl}15:{el}15")
        lc = ws.cell(row=15, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{sl}16:{el}16")
        vc = ws.cell(row=16, column=sc, value=value)
        vc.font = Font(bold=True, size=16, color=WHITE)
        vc.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[15].height = 18
    ws.row_dimensions[16].height = 34

    ws.merge_cells("A17:J17")
    rec_cell = ws["A17"]
    rec_cell.value = (f"  Recommendation: {rec_key}  |  {n_analysts} analysts  |  {upside_str}")
    rec_cell.font = Font(size=10, color=DARK_BLUE, italic=True)
    rec_cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    rec_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[17].height = 20
    ws.row_dimensions[18].height = 10  # spacer

    # ── Outlook & Company Health (rows 19–22, compact) ───────────────────────────
    ws.merge_cells("A19:J19")
    sec19 = ws["A19"]
    sec19.value = "  OUTLOOK & COMPANY HEALTH"
    sec19.font = Font(bold=True, size=11, color=WHITE)
    sec19.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec19.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[19].height = 22

    ws.merge_cells("A20:J22")
    text_cell = ws["A20"]
    text_cell.value = commentary
    text_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left", indent=1)
    text_cell.font = Font(size=10)
    text_cell.fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")
    for r in range(20, 23):
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[23].height = 8  # spacer

    # ── Financial Health Indicators (rows 24–26) ─────────────────────────────────
    ws.merge_cells("A24:J24")
    sec24 = ws["A24"]
    sec24.value = "  FINANCIAL HEALTH INDICATORS"
    sec24.font = Font(bold=True, size=11, color=WHITE)
    sec24.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec24.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[24].height = 22

    dte        = info.get("debtToEquity")
    roe        = info.get("returnOnEquity")
    cr         = info.get("currentRatio")
    qr         = info.get("quickRatio")
    short_pct  = info.get("shortPercentOfFloat")
    short_ratio= info.get("shortRatio")
    ins_pct    = info.get("heldPercentInsiders")
    inst_pct   = info.get("heldPercentInstitutions")

    def _fv(v, mult=1, prefix="", suffix="", decimals=2):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "N/A"
        return f"{prefix}{float(v) * mult:.{decimals}f}{suffix}"

    health_boxes = [
        ("DEBT / EQUITY",    _fv(dte, mult=0.01, suffix="x") if dte else "N/A", "34495E"),
        ("RETURN ON EQUITY", _fv(roe, mult=100, suffix="%", decimals=1),          "1A6B3C"),
        ("CURRENT RATIO",    _fv(cr, suffix="x"),                                 "1F6AA5"),
        ("QUICK RATIO",      _fv(qr, suffix="x"),                                 "0070C0"),
        ("SHORT % FLOAT",    _fv(short_pct, mult=100, suffix="%", decimals=1),    "7B241C"),
    ]
    for idx, (label, value, bg) in enumerate(health_boxes):
        sc = idx * 2 + 1
        sl = chr(ord("A") + sc - 1)
        el = chr(ord("A") + sc)
        ws.merge_cells(f"{sl}25:{el}25")
        lc = ws.cell(row=25, column=sc, value=label)
        lc.font = Font(bold=True, size=9, color=WHITE)
        lc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        lc.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"{sl}26:{el}26")
        vc = ws.cell(row=26, column=sc, value=value)
        vc.font = Font(bold=True, size=14, color=WHITE)
        vc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        vc.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[25].height = 18
    ws.row_dimensions[26].height = 30

    # Ownership / sentiment sub-row (27)
    ws.merge_cells("A27:J27")
    own = ws["A27"]
    parts_own = []
    if ins_pct is not None:
        parts_own.append(f"Insider ownership: {ins_pct * 100:.1f}%")
    if inst_pct is not None:
        parts_own.append(f"Institutional ownership: {inst_pct * 100:.1f}%")
    if short_ratio is not None:
        parts_own.append(f"Short ratio (days to cover): {short_ratio:.1f}")
    own.value = "  " + "   |   ".join(parts_own) if parts_own else "  Ownership data unavailable."
    own.font = Font(size=10, italic=True, color=DARK_BLUE)
    own.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    own.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[27].height = 20

    ws.row_dimensions[28].height = 8  # spacer

    # ── Recent News (rows 29+) ────────────────────────────────────────────────────
    news = news or []
    ws.merge_cells("A29:J29")
    sec29 = ws["A29"]
    sec29.value = "  RECENT NEWS"
    sec29.font = Font(bold=True, size=11, color=WHITE)
    sec29.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec29.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[29].height = 22

    if news:
        for i, item in enumerate(news[:6]):
            r = 30 + i
            bg = ALT_ROW if i % 2 == 0 else WHITE
            date_lbl = f"[{item['date']}]" if item.get("date") else ""
            pub_lbl  = item.get("publisher", "")

            ws.merge_cells(f"A{r}:B{r}")
            meta = ws[f"A{r}"]
            meta.value = f"{date_lbl}  {pub_lbl}"
            meta.font = Font(size=9, italic=True, color="555555")
            meta.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            meta.alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells(f"C{r}:J{r}")
            body = ws[f"C{r}"]
            body.value = item.get("summary", item.get("title", ""))
            body.font = Font(size=10)
            body.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            body.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[r].height = 36
    else:
        ws.merge_cells("A30:J30")
        ws["A30"].value = "  No recent news available."
        ws["A30"].font = Font(size=10, italic=True, color="888888")
        ws.row_dimensions[30].height = 20


# ── Industry peer data ──────────────────────────────────────────────────────────

CELL_GREEN  = "C6EFCE"   # light green fill for "beats average"
CELL_RED    = "FFCCCC"   # light red fill for "trails average"
CELL_NEUT   = "F5F5F5"   # neutral (within ±5%)
TEXT_GREEN  = "276221"
TEXT_RED    = "9C0006"


def _safe_avg(values):
    """Mean of a list, ignoring None and NaN."""
    clean = [v for v in values if v is not None and not (isinstance(v, float) and pd.isna(v))]
    return sum(clean) / len(clean) if clean else None


def _cmp_color(company_val, avg_val, higher_is_better: bool, threshold: float = 0.05):
    """Return (fill_hex, font_hex) based on company vs average."""
    if company_val is None or avg_val is None or avg_val == 0:
        return CELL_NEUT, "000000"
    diff = (company_val - avg_val) / abs(avg_val)
    if abs(diff) <= threshold:
        return CELL_NEUT, "000000"
    beats = diff > 0 if higher_is_better else diff < 0
    return (CELL_GREEN, TEXT_GREEN) if beats else (CELL_RED, TEXT_RED)


def fetch_industry_peers(symbol: str, info: dict, max_peers: int = 12) -> list:
    """
    Return a list of dicts (one per company) with key metrics for all top
    companies in the same yfinance industry as `symbol`.
    The target company is always included and flagged with is_target=True.
    """
    industry_key = info.get("industryKey")
    if not industry_key:
        print("  No industryKey in ticker data — skipping peer comparison.")
        return []

    industry_label = info.get("industry", industry_key)
    print(f"  Fetching industry peers: {industry_label}...")

    try:
        industry_obj = yf.Industry(industry_key)
        top_df = industry_obj.top_companies
    except Exception as exc:
        print(f"  Could not load industry data: {exc}")
        return []

    if top_df is None or (hasattr(top_df, "empty") and top_df.empty):
        print("  No peer data returned.")
        return []

    # Extract ticker list — index is usually the symbol
    try:
        tickers = list(top_df.index)
    except Exception:
        tickers = []

    # Ensure target is present; cap total
    sym_upper = symbol.upper()
    if sym_upper not in [t.upper() for t in tickers]:
        tickers = [sym_upper] + tickers
    tickers = tickers[:max_peers]

    peers = []
    for i, t_sym in enumerate(tickers):
        try:
            d = yf.Ticker(t_sym).info or {}
            name = d.get("longName") or d.get("shortName")
            if not name:
                continue
            peers.append({
                "symbol":           t_sym,
                "name":             name,
                "marketCap":        d.get("marketCap"),
                "trailingPE":       d.get("trailingPE"),
                "forwardPE":        d.get("forwardPE"),
                "grossMargins":     d.get("grossMargins"),
                "profitMargins":    d.get("profitMargins"),
                "revenueGrowth":    d.get("revenueGrowth"),
                "fiftyTwoWeekChange": d.get("52WeekChange"),
                "is_target":        t_sym.upper() == sym_upper,
            })
            print(f"    [{i + 1}/{len(tickers)}] {name}")
        except Exception:
            pass

    return peers


def build_industry_sheet(ws, symbol: str, info: dict, peers: list):
    """
    Three sections:
      1. Industry aggregate stats
      2. Company vs. Industry comparison (color-coded)
      3. Full peer comparison table
    """
    if not peers:
        ws["A1"].value = "Industry peer data not available."
        return

    industry_label = info.get("industry", info.get("industryKey", "Industry"))
    company_name   = info.get("longName", symbol)
    n_peers        = len(peers)

    # ── helpers ──────────────────────────────────────────────────────────────────
    def pct(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "N/A"
        return f"{val * 100:.1f}%"

    def pe(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "N/A"
        return f"{float(val):.1f}x"

    def diff_str(company_val, avg_val):
        if company_val is None or avg_val is None or avg_val == 0:
            return "N/A"
        d = (company_val - avg_val) / abs(avg_val) * 100
        return f"{'+' if d >= 0 else ''}{d:.1f}%"

    # ── aggregate metrics ─────────────────────────────────────────────────────────
    total_mktcap    = sum(p["marketCap"]          for p in peers if p["marketCap"])
    avg_trailing_pe = _safe_avg([p["trailingPE"]       for p in peers])
    avg_forward_pe  = _safe_avg([p["forwardPE"]        for p in peers])
    avg_gross_margin= _safe_avg([p["grossMargins"]     for p in peers])
    avg_net_margin  = _safe_avg([p["profitMargins"]    for p in peers])
    avg_rev_growth  = _safe_avg([p["revenueGrowth"]    for p in peers])
    avg_52wk_chg    = _safe_avg([p["fiftyTwoWeekChange"] for p in peers])

    target = next((p for p in peers if p["is_target"]), peers[0])

    # Column widths
    col_widths = {"A": 30, "B": 16, "C": 16, "D": 16, "E": 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ── Section 1: Header ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    h = ws[f"A{row}"]
    h.value = f"  {industry_label.upper()}  —  Industry Benchmarks"
    h.font = Font(bold=True, size=14, color=WHITE)
    h.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    sub = ws[f"A{row}"]
    sub.value = f"  {n_peers} companies analysed  |  Source: Yahoo Finance  |  {datetime.now().strftime('%B %d, %Y')}"
    sub.font = Font(size=9, italic=True, color=DARK_BLUE)
    sub.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 2  # spacer

    # ── Section 2: Company vs. Industry comparison ───────────────────────────────
    ws.merge_cells(f"A{row}:E{row}")
    sec2 = ws[f"A{row}"]
    sec2.value = f"  {company_name}  vs.  Industry Average"
    sec2.font = Font(bold=True, size=11, color=WHITE)
    sec2.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    # Sub-header
    for col, label in [(1, "Metric"), (2, company_name[:20]), (3, "Industry Avg"),
                       (4, "vs. Avg"), (5, "Signal")]:
        c = ws.cell(row=row, column=col, value=label)
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # Metric definitions: (label, company_value, avg_value, display_fn, higher_is_better)
    metrics = [
        ("Market Cap",        target["marketCap"],        total_mktcap / n_peers,
         fmt_money,                                    True),
        ("P/E Ratio (Trailing)", target["trailingPE"],   avg_trailing_pe,
         pe,                                            False),  # lower = green
        ("P/E Ratio (Forward)",  target["forwardPE"],    avg_forward_pe,
         pe,                                            False),
        ("Gross Margin",      target["grossMargins"],    avg_gross_margin,
         pct,                                           True),
        ("Net Profit Margin", target["profitMargins"],   avg_net_margin,
         pct,                                           True),
        ("Revenue Growth (YoY)", target["revenueGrowth"], avg_rev_growth,
         pct,                                           True),
        ("52-Week Price Change", target["fiftyTwoWeekChange"], avg_52wk_chg,
         pct,                                           True),
    ]

    for i, (label, co_val, avg_val, fmt_fn, higher_better) in enumerate(metrics):
        bg = ALT_ROW if i % 2 == 0 else WHITE
        fill_c, font_c = _cmp_color(co_val, avg_val, higher_better)

        def write(col, value, bold=False, fill=bg, color="000000", align="left"):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(size=10, bold=bold, color=color)
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center")

        write(1, label, bold=True)
        write(2, fmt_fn(co_val),   align="center")
        write(3, fmt_fn(avg_val),  align="center")
        write(4, diff_str(co_val, avg_val), fill=fill_c, color=font_c, align="center")

        # Signal word
        if fill_c == CELL_GREEN:
            sig, sig_color = "Above avg", TEXT_GREEN
        elif fill_c == CELL_RED:
            sig, sig_color = "Below avg", TEXT_RED
        else:
            sig, sig_color = "At par", "666666"
        c5 = ws.cell(row=row, column=5, value=sig)
        c5.font = Font(size=10, bold=True, color=sig_color)
        c5.fill = PatternFill(start_color=fill_c, end_color=fill_c, fill_type="solid")
        c5.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # spacer

    # ── Section 3: Full peer comparison table ────────────────────────────────────
    # Expand columns for wider table
    peer_cols = {"A": 32, "B": 10, "C": 14, "D": 10, "E": 10,
                 "F": 14, "G": 14, "H": 14, "I": 14}
    for col, w in peer_cols.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells(f"A{row}:I{row}")
    sec3 = ws[f"A{row}"]
    sec3.value = "  Full Peer Comparison"
    sec3.font = Font(bold=True, size=11, color=WHITE)
    sec3.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    sec3.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    peer_headers = ["Company", "Ticker", "Market Cap", "P/E Trail",
                    "P/E Fwd", "Gross Margin", "Net Margin", "Rev Growth", "52-Wk Chg"]
    for j, h_label in enumerate(peer_headers, start=1):
        c = ws.cell(row=row, column=j, value=h_label)
        c.font = Font(bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        c.alignment = Alignment(horizontal="center" if j > 1 else "left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    peers_sorted = sorted(peers, key=lambda p: p["marketCap"] or 0, reverse=True)

    # Pre-compute max and min for each metric column (cols 3-9)
    metric_keys = ["marketCap", "trailingPE", "forwardPE",
                   "grossMargins", "profitMargins", "revenueGrowth", "fiftyTwoWeekChange"]
    metric_col  = {k: j for j, k in enumerate(metric_keys, start=3)}  # key -> col number

    def _clean(vals):
        return [v for v in vals if v is not None and not (isinstance(v, float) and pd.isna(v))]

    col_max = {k: max(_clean([p[k] for p in peers_sorted]), default=None) for k in metric_keys}
    col_min = {k: min(_clean([p[k] for p in peers_sorted]), default=None) for k in metric_keys}

    data_row_start = row  # remember where peer rows begin

    for i, p in enumerate(peers_sorted):
        is_tgt = p["is_target"]
        row_bg = "D6E4F0" if is_tgt else (ALT_ROW if i % 2 == 0 else WHITE)

        # Company name + ticker (no max/min highlighting)
        for col, val, align in [(1, p["name"][:40], "left"), (2, p["symbol"], "center")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=10, bold=is_tgt, color=DARK_BLUE if is_tgt else "000000")
            c.fill = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            c.alignment = Alignment(horizontal=align, vertical="center")

        # Metric columns — apply max/min highlight on top of row background
        display_fns = {
            "marketCap":         fmt_money,
            "trailingPE":        pe,
            "forwardPE":         pe,
            "grossMargins":      pct,
            "profitMargins":     pct,
            "revenueGrowth":     pct,
            "fiftyTwoWeekChange": pct,
        }
        for key in metric_keys:
            col = metric_col[key]
            raw = p[key]
            display = display_fns[key](raw)
            is_max = (raw is not None and col_max[key] is not None and
                      not (isinstance(raw, float) and pd.isna(raw)) and raw == col_max[key])
            is_min = (raw is not None and col_min[key] is not None and
                      not (isinstance(raw, float) and pd.isna(raw)) and raw == col_min[key])

            if is_max:
                cell_bg, cell_fc = CELL_GREEN, TEXT_GREEN
            elif is_min:
                cell_bg, cell_fc = CELL_RED, TEXT_RED
            else:
                cell_bg, cell_fc = row_bg, (DARK_BLUE if is_tgt else "000000")

            c = ws.cell(row=row, column=col, value=display)
            c.font = Font(size=10, bold=(is_tgt or is_max or is_min), color=cell_fc)
            c.fill = PatternFill(start_color=cell_bg, end_color=cell_bg, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 18
        row += 1

    ws.merge_cells(f"A{row}:I{row}")
    note = ws[f"A{row}"]
    note.value = ("  * Target company highlighted in blue.  "
                  "Green cell = highest in column.  Red cell = lowest in column.  "
                  "All data from Yahoo Finance.")
    note.font = Font(size=8, italic=True, color="888888")
    ws.row_dimensions[row].height = 14


# ── Name → Ticker lookup ────────────────────────────────────────────────────────

def resolve_ticker(query: str) -> str:
    """
    Accept a company name or ticker and return a confirmed ticker symbol.
    - If the query is a short, space-free string, try it as a ticker first.
    - Otherwise (or if the direct lookup fails) search by name and let the
      user pick from up to 5 equity results.
    """
    query = query.strip()

    # Fast path: looks like a ticker already (e.g. AAPL, BRK.B, META)
    looks_like_ticker = len(query) <= 6 and " " not in query
    if looks_like_ticker:
        test = yf.Ticker(query.upper())
        if test.info and test.info.get("longName"):
            return query.upper()

    # Name search
    print(f"Searching for '{query}'...")
    try:
        results = yf.Search(query, max_results=8, news_count=0)
        quotes = [q for q in results.quotes if q.get("quoteType") == "EQUITY"]
    except Exception:
        quotes = []

    if not quotes:
        # Fall back to using the input as-is
        print(f"No search results found. Trying '{query.upper()}' as a ticker directly.")
        return query.upper()

    # If the raw input exactly matches one of the returned symbols, auto-select it
    upper_query = query.upper()
    for q in quotes:
        if q.get("symbol", "").upper() == upper_query:
            return upper_query

    # Present a numbered list
    display = quotes[:5]
    print("\n  # Company                              Ticker   Exchange")
    print("  " + "-" * 58)
    for i, q in enumerate(display, start=1):
        name = (q.get("longname") or q.get("shortname") or "N/A")[:38]
        sym = q.get("symbol", "N/A")
        exch = q.get("exchange", "")
        print(f"  {i}. {name:<38} {sym:<8} {exch}")

    while True:
        choice = input("\nSelect a number (or press Enter for #1): ").strip()
        if choice == "":
            return display[0]["symbol"]
        if choice.isdigit() and 1 <= int(choice) <= len(display):
            return display[int(choice) - 1]["symbol"]
        print("  Invalid — enter a number from the list above.")


# ── Main ────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        raw_input = " ".join(sys.argv[1:])   # allow "python script.py Apple Inc"
    else:
        raw_input = input("Enter company name or ticker (e.g. Apple or AAPL): ").strip()

    if not raw_input:
        print("Error: no input provided.")
        sys.exit(1)

    symbol = resolve_ticker(raw_input)

    if not symbol:
        print("Could not resolve a ticker. Exiting.")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f"{symbol}_financials.xlsx")

    if os.path.exists(output_path):
        ans = input(f"\n{output_path} already exists. Overwrite? (y/N): ").strip().lower()
        if ans != "y":
            print("Aborted.")
            sys.exit(0)

    # Fetch
    try:
        _ticker, info, quarterly_income, history = fetch_data(symbol)
    except Exception as exc:
        print(f"Error fetching data: {exc}")
        sys.exit(1)

    if not info.get("longName"):
        print(f"No data found for ticker '{symbol}'. Check the symbol and try again.")
        sys.exit(1)

    company_name = info.get("longName", symbol)

    # Generate commentary (AI if key present, rule-based otherwise)
    print("\nBuilding report...")
    news = fetch_recent_news(symbol)
    print(f"  Fetched {len(news)} recent news items.")
    news = generate_news_summaries(news, company_name)
    commentary = generate_ai_commentary(info, quarterly_income, history, news=news)

    # Fetch industry peers (adds ~15-20 seconds)
    peers = fetch_industry_peers(symbol, info)

    # CLI summary
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", "N/A")
    print(f"\n{'=' * 62}")
    print(f"  {company_name}  ({symbol})")
    print(f"{'=' * 62}")
    print(f"  Market Cap   : {fmt_money(info.get('marketCap'))}")
    print(f"  Price        : ${current_price}")
    print(f"  P/E (trail.) : {fmt_val(info.get('trailingPE'), suffix='x')}")
    print(f"  52-Wk Range  : ${info.get('fiftyTwoWeekLow', 'N/A')} - ${info.get('fiftyTwoWeekHigh', 'N/A')}")
    print(f"  Gross Margin : {fmt_val(info.get('grossMargins', 0) * 100 if info.get('grossMargins') else None, suffix='%', decimals=1)}")
    target_mean = info.get("targetMeanPrice")
    rec = (info.get("recommendationKey") or "N/A").upper()
    if target_mean:
        print(f"  Analyst Target: ${target_mean:.2f}  ({rec})")
    print(f"\n  Outlook\n  {'-' * 56}")
    words = commentary.split()
    line, lines = [], []
    for w in words:
        if len(" ".join(line + [w])) > 58:
            lines.append("  " + " ".join(line))
            line = [w]
        else:
            line.append(w)
    if line:
        lines.append("  " + " ".join(line))
    print("\n".join(lines))
    print(f"{'=' * 62}\n")

    # Build Excel — tab order: Dashboard, Industry Comparison, Overview, Income Statement, Revenue Trend
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    build_dashboard_sheet(ws_dash, info, quarterly_income, history, commentary, news=news)

    ws_industry = wb.create_sheet("Industry Comparison")
    build_industry_sheet(ws_industry, symbol, info, peers)

    ws_overview = wb.create_sheet("Overview")
    build_overview_sheet(ws_overview, info)

    ws_income = wb.create_sheet("Income Statement")
    build_income_sheet(ws_income, quarterly_income)

    ws_trend = wb.create_sheet("Revenue Trend")
    build_revenue_trend_sheet(ws_trend, quarterly_income)

    wb.save(output_path)
    print(f"Excel report saved: {output_path}")


if __name__ == "__main__":
    main()
